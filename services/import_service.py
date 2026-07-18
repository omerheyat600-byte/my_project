"""
Import service — brings data from an Excel file into Students, Classes,
Teachers, or Fees.

Design principle: this NEVER writes to the database directly. Every row
is handed to the SAME service method a human would trigger by filling in
the Add Student / Add Class / Add Fee form (student_service.create_student,
class_service.create_class, fee_service.create_fees, ...). That means an
imported row gets every business rule a manually-entered one gets for
free — admission-number auto-assignment, duplicate-class-name checks,
fee status calculation, and (for fees) the automatic Accounts ledger
posting — with zero duplicated logic here to keep in sync.

Row-level, not batch-level: one bad row (missing name, unknown class,
etc.) is recorded as an error and skipped; the rest of the file still
imports. Nothing is rolled back on a partial failure, so re-running an
import after fixing just the failed rows is always safe — already-created
records are naturally skipped as duplicates on the second pass (matched
by admission_no for students, class_name for classes/teachers by name+
subject, student+month+fee_type for fees).

"From another database" workflow: this doesn't connect to a live external
database directly (that would mean handling arbitrary DB engines/creds —
out of scope and risky). The supported path is: export the old system's
data to Excel/CSV (every DB tool and even phpMyAdmin/Access/Google
Sheets can do this), then import that file here using the templates this
service generates.
"""
import os
from datetime import datetime, date

import openpyxl
from openpyxl.styles import Font, PatternFill

from services.student_service import StudentService, StudentValidationError
from services.class_service import ClassService, ClassValidationError, DuplicateClassNameError
from services.teacher_service import TeacherService, TeacherValidationError
from services.fee_service import FeeService, FeeValidationError
from repositories.student_repository import StudentRepository
from utils.logger import get_logger

logger = get_logger(__name__)


class ImportEntityNotSupportedError(Exception):
    pass


# Column headers (normalized: lowercased, spaces/hyphens -> underscore)
# expected for each entity's template, in display order. Marked columns
# are required; the rest are optional and safely left blank.
TEMPLATES = {
    "students": {
        "label": "Students",
        "columns": [
            ("name", True, "Ali Raza"),
            ("grade", True, "Grade 6"),
            ("gender", False, "Male"),
            ("dob", False, "2013-05-14"),
            ("phone", False, "03001234567"),
            ("email", False, "ali.raza@example.com"),
            ("address", False, "House 12, Street 4, Lahore"),
            ("parent_name", False, "Raza Ahmed"),
            ("parent_phone", False, "03007654321"),
            ("join_date", False, "2026-01-15"),
            ("admission_no", False, "2026-0001  (leave blank to auto-assign)"),
            ("roll_no", False, "12"),
        ],
    },
    "classes": {
        "label": "Classes",
        "columns": [
            ("class_name", True, "Grade 6 - A"),
            ("grade_level", True, "Grade 6"),
            ("section", False, "A"),
            ("room_number", False, "12"),
            ("capacity", False, "35"),
            ("max_subjects", False, "8"),
        ],
    },
    "teachers": {
        "label": "Teachers",
        "columns": [
            ("name", True, "Ayesha Khan"),
            ("subject", False, "Mathematics"),
            ("phone", False, "03001234567"),
            ("email", False, "ayesha.khan@example.com"),
            ("address", False, "Model Town, Lahore"),
            ("qualification", False, "M.Sc Mathematics"),
            ("join_date", False, "2024-08-01"),
            ("salary", False, "45000"),
        ],
    },
    "fees": {
        "label": "Fees",
        "columns": [
            ("student_admission_no", True, "2026-0001  (or use student_id if re-importing from this system)"),
            ("fee_type", True, "Tuition Fee"),
            ("month", True, "January"),
            ("amount", True, "5000"),
            ("paid_amount", False, "5000  (0 if unpaid)"),
            ("due_date", False, "2026-01-10"),
            ("paid_date", False, "2026-01-08"),
            ("discount_amount", False, "0"),
            ("discount_reason", False, ""),
            ("fine_amount", False, "0"),
            ("payment_method", False, "Cash"),
        ],
    },
}


class ImportService:

    def __init__(self, student_service: StudentService, class_service: ClassService,
                 teacher_service: TeacherService, fee_service: FeeService,
                 student_repository: StudentRepository = None):
        self.student_service = student_service
        self.class_service = class_service
        self.teacher_service = teacher_service
        self.fee_service = fee_service
        self.student_repository = student_repository or StudentRepository()

    # ---------- Template generation ----------

    def generate_template(self, entity):
        spec = TEMPLATES.get(entity)
        if not spec:
            raise ImportEntityNotSupportedError(f"Unknown import type: {entity}")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = spec["label"][:31]

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
        example_font = Font(italic=True, color="6B7280")

        for col_idx, (name, required, _example) in enumerate(spec["columns"], start=1):
            cell = ws.cell(row=1, column=col_idx, value=name + (" *" if required else ""))
            cell.font = header_font
            cell.fill = header_fill
            ws.column_dimensions[cell.column_letter].width = max(18, len(name) + 4)

        for col_idx, (_name, _required, example) in enumerate(spec["columns"], start=1):
            cell = ws.cell(row=2, column=col_idx, value=example)
            cell.font = example_font

        # Legend row explaining how to use the template, a few rows below
        # the example so it doesn't get mistaken for data.
        legend_row = 4
        ws.cell(row=legend_row, column=1,
                value="Columns marked * are required. Row 2 is an example — replace/delete it before importing.").font = Font(italic=True, size=10, color="6B7280")
        if entity == "fees":
            ws.cell(row=legend_row + 1, column=1,
                    value="student_admission_no must match an existing student already imported/created in this system.").font = Font(italic=True, size=10, color="6B7280")

        return wb

    # ---------- Reading uploaded files ----------

    def _read_rows(self, file_path):
        """Reads the first sheet of an .xlsx file into a list of dicts,
        keyed by normalized header (lowercased, spaces/hyphens -> _).
        Skips the example/legend rows this service's own templates add,
        and any fully-blank row."""
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active

        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return []

        headers = [
            str(h).strip().lower().replace(" ", "_").replace("-", "_").rstrip("*").rstrip("_")
            if h else ""
            for h in header_row
        ]

        rows = []
        for raw in rows_iter:
            if raw is None or all(v is None or str(v).strip() == "" for v in raw):
                continue
            row = {}
            for idx, header in enumerate(headers):
                if not header:
                    continue
                value = raw[idx] if idx < len(raw) else None
                if isinstance(value, datetime):
                    value = value.date().isoformat()
                elif isinstance(value, date):
                    value = value.isoformat()
                elif value is not None:
                    value = str(value).strip()
                row[header] = value if value not in (None, "") else None
            # Skip rows that are just this service's own italic example
            # row wording leaking through (e.g. "(leave blank to auto-assign)").
            if any(v and "leave blank" in str(v).lower() for v in row.values()):
                continue
            rows.append(row)
        return rows

    # ---------- Students ----------

    def import_students(self, file_path, created_by=None):
        rows = self._read_rows(file_path)
        imported, skipped, errors = 0, 0, []

        for i, row in enumerate(rows, start=2):  # row 1 = header
            admission_no = (row.get("admission_no") or "").strip() if row.get("admission_no") else ""
            try:
                if admission_no and self.student_repository.admission_no_exists(admission_no):
                    skipped += 1
                    errors.append({"row": i, "reason": f"Skipped — admission_no '{admission_no}' already exists"})
                    continue
                self.student_service.create_student(row)
                imported += 1
            except StudentValidationError as e:
                errors.append({"row": i, "reason": "; ".join(e.errors)})
            except Exception as e:
                logger.warning(f"Import students: row {i} failed: {e}")
                errors.append({"row": i, "reason": str(e)})

        logger.info(f"Import students: {imported} imported, {skipped} skipped, {len(errors)} errors (by {created_by})")
        return {"total_rows": len(rows), "imported": imported, "skipped": skipped, "errors": errors}

    # ---------- Classes ----------

    def import_classes(self, file_path, created_by=None):
        rows = self._read_rows(file_path)
        imported, skipped, errors = 0, 0, []

        for i, row in enumerate(rows, start=2):
            try:
                self.class_service.create_class(row)
                imported += 1
            except DuplicateClassNameError:
                skipped += 1
                errors.append({"row": i, "reason": f"Skipped — class '{row.get('class_name')}' already exists"})
            except ClassValidationError as e:
                errors.append({"row": i, "reason": "; ".join(e.errors)})
            except Exception as e:
                logger.warning(f"Import classes: row {i} failed: {e}")
                errors.append({"row": i, "reason": str(e)})

        logger.info(f"Import classes: {imported} imported, {skipped} skipped, {len(errors)} errors (by {created_by})")
        return {"total_rows": len(rows), "imported": imported, "skipped": skipped, "errors": errors}

    # ---------- Teachers ----------

    def import_teachers(self, file_path, created_by=None):
        rows = self._read_rows(file_path)
        imported, skipped, errors = 0, 0, []

        for i, row in enumerate(rows, start=2):
            try:
                self.teacher_service.create_teacher(row)
                imported += 1
            except TeacherValidationError as e:
                errors.append({"row": i, "reason": "; ".join(e.errors)})
            except Exception as e:
                logger.warning(f"Import teachers: row {i} failed: {e}")
                errors.append({"row": i, "reason": str(e)})

        logger.info(f"Import teachers: {imported} imported, {skipped} skipped, {len(errors)} errors (by {created_by})")
        return {"total_rows": len(rows), "imported": imported, "skipped": skipped, "errors": errors}

    # ---------- Fees ----------

    def import_fees(self, file_path, created_by=None):
        """
        Each row goes through fee_service.create_fees exactly like a
        manually-entered fee would — so an imported row with a nonzero
        paid_amount automatically gets posted to the Accounts ledger too
        (via FeeAccountingService), same as if staff had typed it in.
        """
        rows = self._read_rows(file_path)
        imported, skipped, errors = 0, 0, []

        for i, row in enumerate(rows, start=2):
            student_id = (row.get("student_id") or "").strip() if row.get("student_id") else ""
            admission_no = (row.get("student_admission_no") or row.get("admission_no") or "").strip() \
                if (row.get("student_admission_no") or row.get("admission_no")) else ""

            resolved_id = None
            if student_id and self.student_repository.find_by_id(student_id):
                resolved_id = student_id
            elif admission_no:
                student = self.student_repository.find_by_admission_no(admission_no)
                if student:
                    resolved_id = student.id

            if not resolved_id:
                skipped += 1
                identifier = student_id or admission_no or "(none given)"
                errors.append({"row": i, "reason": f"Skipped — no matching student for '{identifier}'"})
                continue

            fee_data = {**row, "student_id": resolved_id}
            try:
                self.fee_service.create_fees(fee_data, created_by=created_by)
                imported += 1
            except FeeValidationError as e:
                errors.append({"row": i, "reason": "; ".join(e.errors)})
            except Exception as e:
                logger.warning(f"Import fees: row {i} failed: {e}")
                errors.append({"row": i, "reason": str(e)})

        logger.info(f"Import fees: {imported} imported, {skipped} skipped, {len(errors)} errors (by {created_by})")
        return {"total_rows": len(rows), "imported": imported, "skipped": skipped, "errors": errors}

    # ---------- Dispatch ----------

    def run_import(self, entity, file_path, created_by=None):
        handlers = {
            "students": self.import_students,
            "classes": self.import_classes,
            "teachers": self.import_teachers,
            "fees": self.import_fees,
        }
        handler = handlers.get(entity)
        if not handler:
            raise ImportEntityNotSupportedError(f"Unknown import type: {entity}")
        return handler(file_path, created_by)
